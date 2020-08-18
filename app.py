# To add a new cell, type '# %%'
# To add a new markdown cell, type '# %% [markdown]'
# %%
import os
import re
import requests
import torch
import imgurfile
from datetime import date, datetime
import numpy as np
import PIL.Image as Image
import PIL.ImageDraw as ImageDraw
import PIL.ImageFont as ImageFont
from urllib.parse import parse_qs
import csv
import openpyxl as op
import random
import threading
import time
# import imgurfile
from imgurpython import ImgurClient
from linebot.models.template import *
from linebot.models.template import (
    ButtonsTemplate, CarouselTemplate, ConfirmTemplate, ImageCarouselTemplate
)
from linebot.models import (
    PostbackEvent
)
from linebot.models import (
    FollowEvent
)
from linebot.models import (
    ImagemapSendMessage, TextSendMessage, ImageSendMessage, LocationSendMessage, FlexSendMessage, VideoSendMessage
)
import json
from linebot.exceptions import (
    InvalidSignatureError
)
from linebot import (
    LineBotApi, WebhookHandler
)
from linebot.models import (
    MessageEvent, TextMessage, ImageMessage
)
from flask import Flask, request, abort
from IPython import get_ipython

# %%
# get_ipython().system('pip install flask')
# get_ipython().system('pip install line-bot-sdk')


# %%
'''

整體功能描述

應用伺服器主結構樣板
    用來定義使用者傳送消息時，應該傳到哪些方法上
        比如收到文字消息時，轉送到文字專屬處理方法

消息專屬方法定義
    當收到文字消息，從資料夾內提取消息，並進行回傳。

啟動應用伺服器

'''


# %%
'''

Application 主架構

'''

# 引用Web Server套件

# 從linebot 套件包裡引用 LineBotApi 與 WebhookHandler 類別

# 引用無效簽章錯誤

# 載入json處理套件

# 載入基礎設定檔
secretFileContentJson = json.load(
    open("./line_secret_key", 'r', encoding='utf8'))

# 設定Server啟用細節
app = Flask(__name__, static_url_path="/material", static_folder="./material/")

# 生成實體物件
line_bot_api = LineBotApi(secretFileContentJson.get("channel_access_token"))
handler = WebhookHandler(secretFileContentJson.get("secret_key"))

# 啟動server對外接口，使Line能丟消息進來


@app.route("/", methods=['POST'])
def callback():
    # get X-Line-Signature header value
    signature = request.headers['X-Line-Signature']

    # get request body as text
    body = request.get_data(as_text=True)
    app.logger.info("Request body: " + body)

    # handle webhook body
    try:
        handler.handle(body, signature)
    except InvalidSignatureError:
        abort(400)

    return 'OK'


# %%
'''

消息判斷器

讀取指定的json檔案後，把json解析成不同格式的SendMessage

讀取檔案，
把內容轉換成json
將json轉換成消息
放回array中，並把array傳出。

'''

# 引用會用到的套件


def detect_json_array_to_new_message_array(fileName):

    # 開啟檔案，轉成json
    with open(fileName, 'r', encoding='utf8') as f:
        jsonArray = json.load(f)

    # 解析json
    returnArray = []
    for jsonObject in jsonArray:

        # 讀取其用來判斷的元件
        message_type = jsonObject.get('type')

        # 轉換
        if message_type == 'text':
            returnArray.append(TextSendMessage.new_from_json_dict(jsonObject))
        elif message_type == 'imagemap':
            returnArray.append(
                ImagemapSendMessage.new_from_json_dict(jsonObject))
        elif message_type == 'template':
            returnArray.append(
                TemplateSendMessage.new_from_json_dict(jsonObject))
        elif message_type == 'image':
            returnArray.append(ImageSendMessage.new_from_json_dict(jsonObject))
        elif message_type == 'sticker':
            returnArray.append(
                StickerSendMessage.new_from_json_dict(jsonObject))
        elif message_type == 'audio':
            returnArray.append(AudioSendMessage.new_from_json_dict(jsonObject))
        elif message_type == 'location':
            returnArray.append(
                LocationSendMessage.new_from_json_dict(jsonObject))
        elif message_type == 'flex':
            returnArray.append(FlexSendMessage.new_from_json_dict(jsonObject))
        elif message_type == 'video':
            returnArray.append(VideoSendMessage.new_from_json_dict(jsonObject))

    # 回傳
    return returnArray


# %%
'''

handler處理關注消息

用戶關注時，讀取 素材 -> 關注 -> reply.json

將其轉換成可寄發的消息，傳回給Line

'''

# 引用套件

# 關注事件處理

# %%
'''

handler處理文字消息

收到用戶回應的文字消息，
按文字消息內容，往素材資料夾中，找尋以該內容命名的資料夾，讀取裡面的reply.json

轉譯json後，將消息回傳給用戶

'''

# 引用套件


def uploadImage(filename, path):
    config = {
        'name': filename.split('.')[0],
        'title': path.split('.')[0],
        'description': 'test-description'
    }
    return imgurfile.upload(imgur_client, path, config)["link"]


@handler.add(MessageEvent, message=ImageMessage)
def process_image_message(event):
    global userlist
    result_message_array = []
    response = requests.get(
        f"https://api.line.me/v2/bot/message/{event.message.id}/content", stream=True, headers={'Authorization': f'Bearer {secretFileContentJson.get("channel_access_token")}'})

    img = Image.open(response.raw)
    filepath = f"predection/{event.message.id}.{img.format.lower()}"
    img.save(filepath)
    usid = event.source.user_id.replace("\n", "")
    username = [i for i in userlist if i[0].replace(
                "\n", "") == usid]
    username = username[0][1]
    imageweburl = "非管理員無法上傳"
    if username == "陳皇宇" or username == "張文耀":
        imageweburl = (uploadImage(filepath.split("/")[-1], filepath))
    result_message_array.append(TextSendMessage(
        text=f"好歐收到囉~\n\nURL:{imageweburl}"))
    line_bot_api.reply_message(
        event.reply_token,
        result_message_array
    )


def updatemax():
    global usermax, dinnermax, dinnerlist, userlist, usersheet, dinnersheet
    usermax = usersheet.max_row
    dinnermax = dinnersheet.max_row
    dinnerlist = [[b.value for b in i] for i in list(dinnersheet.rows)]
    userlist = [[b.value for b in i] for i in list(usersheet.rows)]


def reminduser(remindid, message):
    push = input("Remind? yes ot n\n")
    if push.lower() == "yes":
        line_bot_api.multicast(remindid, TextSendMessage(text=message))
        # line_bot_api.multicast(notorderid, TextSendMessage(
        #     text='緊急事件，抱歉今天有點晚，請在1530前點完餐，互相提醒發揮愛<3'))
        print(f"已送出提醒！共{len(remindid)}人")


def getAlluser():
    user_Name = [i[1] for i in userlist]
    dinnerlisttmp = []
    for i in range(1, len(dinnerlist)):
        dinnerlisttmp.append(dinnerlist[i][3])
    usernotorder = []
    usernotordername = []
    for i in range(1, len(user_Name)):
        usernotorder.append(userlist[i][0])
        usernotordername.append(userlist[i][1])
    print(usernotordername)
    print(f"共{len(usernotordername)}人")
    return usernotorder


def getNotOrder(printname=True):
    global notordernum
    user_Name = [i[1] for i in userlist]
    dinnerlisttmp = []
    for i in range(1, len(dinnerlist)):
        if dinnerlist[i][0].date() == date.today():
            dinnerlisttmp.append(dinnerlist[i][3])
    usernotorder = []
    usernotordername = []
    for i in range(1, len(user_Name)):
        if user_Name[i] not in dinnerlisttmp:
            usernotorder.append(userlist[i][0])
            usernotordername.append(userlist[i][1])
    if printname:
        print(usernotordername)
    notordernum = len(usernotordername)
    print(f"共{len(usernotordername)}人尚未點餐")
    return usernotorder


@handler.add(MessageEvent, message=TextMessage)
def process_text_message(event):
    global filepath, today_date, storeName
    userinfo = {}
    userinfo["message"] = event.message.text
    userinfo["ID"] = event.source.user_id.replace("\n", "")
    userinfo["token"] = event.reply_token
    userRequireList.append(userinfo)


def randomsuer(number):
    whilenumber = number
    useridlist = []
    usernamelist = []
    while(whilenumber > 0):
        userindex = random.randint(1, len(userlist)-1)
        if userlist[userindex][0] not in useridlist:
            useridlist.append(userlist[userindex][0])
            usernamelist.append(
                userlist[userindex][1])
            whilenumber -= 1
    return useridlist, usernamelist


def Sendtouser(useridlist):
    line_bot_api.multicast(
        useridlist,
        [TextSendMessage(text="恭喜你成為 主線任務 倒垃圾值日生 的幸運兒\n請立刻找執行長確認任務內容")]
    )
    print("Finish")


def writeExcel():
    global userRequireList, app, notordernum
    while(1):
        while(userRequireList != []):
            userinfo = userRequireList.pop(0)
            usermessage = userinfo["message"]
            user_id = userinfo["ID"]
            if usermessage == "Hello, world":
                print("Success!")
                continue
            if usermessage == "格式":
                continue
            if "吃什麼" in usermessage or "今天吃" in usermessage or "現在吃" in usermessage or "晚餐吃" in usermessage or "菜單" in usermessage or usermessage == "晚餐":
                line_bot_api.reply_message(
                    userinfo["token"],
                    getFoodMessage()
                )
                continue
            user_info = [i for i in userlist if i[0].replace(
                "\n", "") == user_id]
            user_index = -1
            if user_info[0][1] == "陳皇宇" and usermessage == "shutdown":
                return
            for index, i in enumerate(userlist):
                if i[0].replace("\n", "") == user_id:
                    user_index = index
                    break
            if user_info == []:
                usersheet[f"A{usermax+1}"] = user_id
                usersheet[f"B{usermax+1}"] = usermessage
                wb.save(filepath)
                updatemax()
                line_bot_api.reply_message(
                    userinfo["token"],
                    TextSendMessage(text=f"{user_id} {usermessage} 已登錄")
                )
                print(f"已登錄：{usermessage}/{user_id}")
            else:
                user_info = user_info[0]
                testinfo = usermessage.replace("／", "/").split("/")
                if len(testinfo) != 3:
                    if usersheet[f"D{user_index+1}"].value != None:
                        usersheet[f"D{user_index+1}"] = usersheet[f"D{user_index+1}"].value + \
                            usermessage+"\n"
                    else:
                        usersheet[f"D{user_index+1}"] = usermessage+"\n"
                    wb.save(filepath)
                    updatemax()
                    print(user_info[1] + ":"+usermessage)
                    if "麥當勞" in usermessage:
                        line_bot_api.reply_message(
                            userinfo["token"],
                            TextSendMessage(text="再麥當勞給我試試看")
                        )
                    elif "歐姐" in usermessage or "歐姊" in usermessage:
                        line_bot_api.reply_message(
                            userinfo["token"],
                            TextSendMessage(text="沒被打過?")
                        )
                    else:
                        line_bot_api.reply_message(
                            userinfo["token"],
                            TextSendMessage(text="願望小精靈收到囉~\n但不一定會實現呦<3")
                        )
                    continue
                check = "F"
                for i in range(1, len(dinnerlist)):
                    if (type(dinnerlist[i][0]) != type(date.today())):
                        dinnerlist[i][0] = dinnerlist[i][0].date()
                    if (dinnerlist[i][0] == date.today() and user_info[1] == dinnerlist[i][3] and storeName == dinnerlist[i][2]):
                        check = "T"
                nowtime = datetime.now().strftime("%H:%M")
                dinnersheet[f"A{dinnermax+1}"] = today_date
                dinnersheet[f"B{dinnermax+1}"] = nowtime
                dinnersheet[f"C{dinnermax+1}"] = storeName
                dinnersheet[f"D{dinnermax+1}"] = user_info[1]
                dinnersheet[f"E{dinnermax+1}"] = testinfo[0]
                dinnersheet[f"F{dinnermax+1}"] = int(testinfo[1])
                dinnersheet[f"G{dinnermax+1}"] = int(testinfo[2])
                dinnersheet[f"H{dinnermax+1}"] = check
                dinnersheet[f"I{dinnermax+1}"] = int(
                    testinfo[1]) - int(testinfo[2])
                wb.save(filepath)
                notordernum -= 1
                updatemax()
                line_bot_api.reply_message(
                    userinfo["token"],
                    TextSendMessage(text=f"{user_info[1]} 飢餓精靈收到囉！")
                )
                print(
                    f"點餐：{today_date}/{nowtime}/{storeName}/{user_info[1]}/{testinfo[0]}/{testinfo[1]}/{testinfo[2]}")
                print(f"尚未點餐人數：{notordernum}")
        time.sleep(0.3)


def getFoodMessage():
    global storeName
    imagepath1 = ""
    imagepath2 = ""
    subsidy = 100
    storeName = "胖老爹"
    headerstring = "【真的最後一次推飯】\n\n菜單： https://www.foodbooking.com/api/fb/eplv6"
    notenough = ""
    formatstring = "餐點/價格/價差(>=0)\n不吃不點：今天不吃/0/0"
    footstring = "都沒玩到桌游真的很可惜\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n天下沒有不散的宴席\n\n\n\n\n\n\n\n\n\n\n\n\n\n"
    FoodMessaage = []
    if imagepath1 != "":
        FoodMessaage.append(ImageSendMessage(
            original_content_url=imagepath1, preview_image_url=imagepath1))
    if imagepath2 != "":
        FoodMessaage.append(ImageSendMessage(
            original_content_url=imagepath2, preview_image_url=imagepath2))
    FoodMessaage.append(TextSendMessage(
        text=f"{headerstring}\n\n店家：{storeName}\n補助：{subsidy}\n缺貨：{notenough}\n{formatstring}\n\n{footstring}"))
    return FoodMessaage


# %%
'''

Application 運行（開發版）

'''
storeName = ""
filepath = "DinnerList.xlsx"
wb = op.load_workbook(filepath)
usersheet = wb["UserList"]
dinnersheet = wb["DinnerList"]
usermax = usersheet.max_row
dinnermax = dinnersheet.max_row
dinnerlist = [i[0].value for i in list(dinnersheet.rows)]
userlist = [i[0].value for i in list(usersheet.rows)]
# print(userlist)
today_date = date.today()
userRequireList = []
notordernum = 33
# ImageSendMessage(original_content_url='圖片網址', preview_image_url='圖片網址')

updatemax()
if __name__ == "__main__":
    # imgur_client = imgurfile.setauthorize()
    # for i in range(15):
    #     acc, loss = FSA.train()
    #     acc = round(acc, 4)
    #     loss = round(loss, 4)
    #     print(f"loss:{loss}  acc:{acc}")
    # process_image_message(None)
    imgur_client = imgurfile.setauthorize()
    getFoodMessage()
    reminduser(getNotOrder(), "真●最後一次推飯")
    # reminduser(getAlluser(), "給大家喝個飲料，補個元氣！大家加油！\n\n請打\"菜單\"看詳細資訊")
    # while(input("random again?") == "yes"):
    #     userids, usernames = randomsuer(2)
    #     print(usernames)
    # if userids != []:
    #     Sendtouser(userids)
    threading.Thread(target=writeExcel).start()
    app.run(host='0.0.0.0')


# %%
'''

Application 運行（heroku版）

'''

# if __name__ == "__main__":
#     imgur_client = imgurfile.setauthorize()
#     app.run(host='0.0.0.0', port=os.environ['PORT'])


# %%
